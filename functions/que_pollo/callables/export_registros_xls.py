"""Callable que genera un XLS con los registros filtrados y devuelve URL prefirmada.

Aplica los mismos filtros server-side que el frontend para que el XLS contenga
exactamente los mismos resultados que el listado, sin paginación.
"""

from __future__ import annotations

from datetime import datetime, timedelta, timezone
from io import BytesIO
from typing import Any
from zoneinfo import ZoneInfo

import google.auth
from google.auth.transport import requests as google_requests
from firebase_admin import firestore, storage
from firebase_functions import https_fn
from google.cloud.firestore_v1.base_query import FieldFilter
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


MX_TZ = ZoneInfo("America/Mexico_City")

from .. import firebase_app as _app  # noqa: F401  (asegura initialize_app)


EXPORT_ROW_LIMIT = 3000

COLUMN_HEADERS = [
    ("Fecha y hora registro", "Fecha_registro"),
    ("Campaña", "campania"),
    ("Folio de Ticket", "Folio_ticket"),
    ("Sucursal", "Sucursal_ticket"),
    ("ID de registro", "ID_registro"),
    ("Nombre del participante", "Nombre_registro"),
    ("WhatsApp", "Whatsapp_registro"),
    ("Estado", "Clasificacion_registro"),
    ("Fecha y hora de Ticket", "Fecha_ticket"),
    ("Monto total", "Monto_ticket"),
    ("¿Respondió correcto?", "Trivia_registro"),
]

CAMPANIA_LABELS = {
    "jerseys": "Jerseys",
    "huevo_campeon": "Huevo Campeón",
}

# Sentinela del último carácter Unicode común para prefix matches.
PREFIX_SENTINEL = ""


def _is_admin(uid: str) -> bool:
    db = firestore.client()
    return db.collection("usuarios_admin").document(uid).get().exists


def _start_of_day(value: str) -> datetime:
    return datetime.fromisoformat(value).replace(
        hour=0, minute=0, second=0, microsecond=0, tzinfo=timezone.utc
    )


def _end_of_day(value: str) -> datetime:
    return datetime.fromisoformat(value).replace(
        hour=23, minute=59, second=59, microsecond=999999, tzinfo=timezone.utc
    )


def _apply_filters(query, data: dict[str, Any]):
    if data.get("campania"):
        query = query.where(filter=FieldFilter("campania", "==", data["campania"]))
    if data.get("clasificacion"):
        query = query.where(
            filter=FieldFilter("Clasificacion_registro", "==", data["clasificacion"])
        )
    if data.get("sucursal"):
        query = query.where(filter=FieldFilter("Sucursal_ticket", "==", data["sucursal"]))
    if data.get("fechaRegistroDesde"):
        query = query.where(
            filter=FieldFilter("Fecha_registro", ">=", _start_of_day(data["fechaRegistroDesde"]))
        )
    if data.get("fechaRegistroHasta"):
        query = query.where(
            filter=FieldFilter("Fecha_registro", "<=", _end_of_day(data["fechaRegistroHasta"]))
        )
    if data.get("fechaTicketDesde"):
        query = query.where(
            filter=FieldFilter("Fecha_ticket", ">=", _start_of_day(data["fechaTicketDesde"]))
        )
    if data.get("fechaTicketHasta"):
        query = query.where(
            filter=FieldFilter("Fecha_ticket", "<=", _end_of_day(data["fechaTicketHasta"]))
        )

    # Prefix match para campos de texto.
    def add_prefix(field: str, key: str):
        value = data.get(key)
        if not value:
            return
        nonlocal query
        query = query.where(filter=FieldFilter(field, ">=", value))
        query = query.where(filter=FieldFilter(field, "<=", value + PREFIX_SENTINEL))

    add_prefix("ID_registro", "idRegistro")
    add_prefix("Folio_ticket", "folio")
    add_prefix("Nombre_registro", "nombre")
    add_prefix("Whatsapp_registro", "whatsapp")

    return query.order_by("Fecha_registro", direction=firestore.Query.DESCENDING)


def _format_cell(field: str, value: Any) -> Any:
    if value is None:
        return ""
    if field in ("Fecha_registro", "Fecha_ticket"):
        try:
            return value.astimezone(MX_TZ).strftime("%d/%m/%Y %H:%M")
        except AttributeError:
            return str(value)
    if field == "Trivia_registro":
        return "Sí" if value == "yes" else "No"
    if field == "Clasificacion_registro":
        return "VÁLIDO" if value == "VALIDO" else "INVÁLIDO"
    if field == "campania":
        return CAMPANIA_LABELS.get(value, value)
    return value


def _build_workbook(rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Registros"

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF7D15B4")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, (label, _) in enumerate(COLUMN_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[cell.column_letter].width = max(18, len(label) + 2)

    for row_idx, doc in enumerate(rows, start=2):
        for col_idx, (_, field) in enumerate(COLUMN_HEADERS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_format_cell(field, doc.get(field)))

    ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@https_fn.on_call(region="us-central1", memory=512, max_instances=5)
def export_registros_xls(req: https_fn.CallableRequest) -> dict[str, str]:
    if not req.auth or not req.auth.uid:
        raise https_fn.HttpsError(https_fn.FunctionsErrorCode.UNAUTHENTICATED, "Login requerido")

    uid = req.auth.uid
    if not _is_admin(uid):
        raise https_fn.HttpsError(https_fn.FunctionsErrorCode.PERMISSION_DENIED, "No autorizado")

    data = req.data if isinstance(req.data, dict) else {}

    db = firestore.client()
    query = _apply_filters(db.collection("registros"), data).limit(EXPORT_ROW_LIMIT)
    rows = [doc.to_dict() for doc in query.stream()]
    truncated = len(rows) >= EXPORT_ROW_LIMIT

    payload = _build_workbook(rows)

    bucket = storage.bucket()
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    filename = f"registros-{timestamp}.xlsx"
    blob_path = f"exports/{uid}/{filename}"
    blob = bucket.blob(blob_path)
    blob.upload_from_string(
        payload,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # En Cloud Functions las credenciales por defecto no traen private key,
    # así que firmamos vía IAM signBlob (requiere rol serviceAccountTokenCreator
    # sobre la propia SA del runtime).
    credentials, _ = google.auth.default()
    credentials.refresh(google_requests.Request())
    url = blob.generate_signed_url(
        expiration=timedelta(minutes=10),
        method="GET",
        version="v4",
        service_account_email=credentials.service_account_email,
        access_token=credentials.token,
    )
    return {
        "url": url,
        "filename": filename,
        "count": str(len(rows)),
        "truncated": "true" if truncated else "false",
        "limit": str(EXPORT_ROW_LIMIT),
    }
